"""Qlever + Photon + NCES facility finder — v6.

v6 layers the Urban Institute Education Data API (federal NCES Common Core
of Data) on top of Qlever/OSM. Every US public K-12 school appears even
when OSM has no pitch tagging. NCES supplies name, GPS, and school_level
(Elementary/Middle/High); the memo dimension table supplies default field
sizes when no OSM pitch is present.

Data flow additions vs v5:
1. fetch_nces_schools(city, state) hits
   https://educationdata.urban.org/api/v1/schools/ccd/directory/YEAR/
   Results cached per state for 30 days.
2. fetch_overpass merges NCES entries with Qlever entries; coord dedup
   drops overlaps (OSM school + same NCES school).
3. merge_and_deduplicate lets source == "nces" bypass sport-gating
   post-filter — every real school is a valid facility for the memo sports.
4. categorize routes NCES entries by school_level (1=Elem, 2=Mid, 3=Hi).

Legacy v5 header follows.
--
Qlever + Photon facility finder — v5 (broader container fetch).

v5 drops the SPARQL spatial-join `ogc:sfContains ?pitch` requirement from
the parks/schools/sports_centres/recreation_grounds queries. Earlier
versions required a facility to already contain an OSM-tagged pitch, which
silently excluded middle schools and other facilities whose pitches were
not tagged. v5 fetches all containers in the city bbox, then relies on the
existing 500m haversine proximity step to attach pitches; the
`has_pitches or is_sport_name` post-filter still prevents the
"every park appears in every sport" bug.

Improvements over qlever_facility_finder.py:
1. Container queries (parks/schools/sports_centres) now use Qlever's spatial
   join `?facility ogc:sfContains ?pitch` to find facilities that actually
   contain a pitch of the chosen sport — replaces the 200m haversine
   heuristic which missed large campuses.
2. Indoor sports (Basketball, Volleyball) include ALL sports_centres,
   fitness_centres, sports_halls, and community_centres regardless of
   pitch tags (indoor courts are rarely tagged with sport).
3. Adds athletic_centre, stadium, recreation_ground, gym building tags to
   the candidate facility pool.

Same input/output schema as qlever_facility_finder.py.
"""

import re
import time
import math
import io
import json
import hashlib
import sqlite3
import os
import zipfile
import traceback
import requests
import pandas as pd
import streamlit as st
from dotenv import load_dotenv

load_dotenv()
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OVERPASS_MIRRORS = [
    "https://overpass-api.de/api/interpreter",
    "https://overpass.kumi.systems/api/interpreter",
    "https://overpass.osm.ch/api/interpreter",
    "https://overpass.openstreetmap.fr/api/interpreter",
]
DEFAULT_OVERPASS_URL = OVERPASS_MIRRORS[0]

NOMINATIM_URL = "https://nominatim.openstreetmap.org"

_CONTACT = os.environ.get("CONTACT_EMAIL", "")
if _CONTACT:
    USER_AGENT = f"SportsFacilityFinder/1.0 ({_CONTACT})"
else:
    USER_AGENT = "SportsFacilityFinder/1.0"

HEADERS = {
    "User-Agent": USER_AGENT,
    "Accept": "application/json",
    "Accept-Language": "en-US,en;q=0.9",
}

_log_lock = Lock()

# Cap concurrent SPARQL POSTs to the public Qlever endpoint across all
# worker threads. Batch mode runs N outer x M inner workers = up to N*M
# simultaneous queries; the public instance silently throttles -> returns []
# -> "0 facilities" everywhere. Cap globally at 3.
from threading import Semaphore as _Semaphore
_QLEVER_SEM = _Semaphore(3)

from threading import Semaphore as _Semaphore
_OVERPASS_SEM = _Semaphore(2)
_NOMINATIM_SEM = _Semaphore(1)

CACHE_DB_PATH = "facility_cache.db"
CACHE_TTL_SECONDS = 7 * 24 * 3600

_cache_lock = Lock()

def _init_cache():
    with sqlite3.connect(CACHE_DB_PATH) as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS api_cache (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL,
                created_at INTEGER NOT NULL
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_created ON api_cache(created_at)")

def _cache_key(prefix, *args):
    raw = prefix + "|" + "|".join(str(a) for a in args)
    return hashlib.sha256(raw.encode()).hexdigest()

def cache_get(key, max_age_seconds=None):
    ttl = max_age_seconds if max_age_seconds is not None else CACHE_TTL_SECONDS
    with _cache_lock:
        with sqlite3.connect(CACHE_DB_PATH) as conn:
            row = conn.execute(
                "SELECT value, created_at FROM api_cache WHERE key = ?", (key,)
            ).fetchone()
            if not row:
                return None
            value, created_at = row
            if time.time() - created_at > ttl:
                return None
            try:
                return json.loads(value)
            except Exception:
                return None

def cache_set(key, value):
    with _cache_lock:
        with sqlite3.connect(CACHE_DB_PATH) as conn:
            conn.execute(
                "INSERT OR REPLACE INTO api_cache (key, value, created_at) VALUES (?, ?, ?)",
                (key, json.dumps(value), int(time.time())),
            )

def cache_clear():
    with _cache_lock:
        with sqlite3.connect(CACHE_DB_PATH) as conn:
            conn.execute("DELETE FROM api_cache")

def cache_stats():
    if not os.path.exists(CACHE_DB_PATH):
        return 0, 0
    with _cache_lock:
        with sqlite3.connect(CACHE_DB_PATH) as conn:
            count = conn.execute("SELECT COUNT(*) FROM api_cache").fetchone()[0]
    size = os.path.getsize(CACHE_DB_PATH)
    return count, size

_init_cache()

SPORTS_CONFIG = {
    "Soccer / Football": {
        "osm_sports": ["soccer", "football"],
        "keywords": ["soccer", "football", "futbol", "fútbol", "athletic field",
                     "sports field", "multi-purpose", "multipurpose"],
        "exclude": ["swim", "pool", "aqua", "skatepark", "golf", "bowling",
                    "tennis center", "library", "marina", "model airplane"],
        "facility_label": "Soccer Field",
        "label_variants": {
            "soccer_football": "Soccer/Football Field",
            "football_only": "Football Field",
            "multi": "Multi-Purpose Field (Soccer)",
        },
        "section_keywords": ["park", "field", "sports", "recreation"],
        "min_pitch_length_ft": 180,
    },
    "Baseball / Softball": {
        "osm_sports": ["baseball", "softball"],
        "keywords": ["baseball", "softball", "diamond", "little league",
                     "ball field", "ballfield", "tee ball", "t-ball"],
        "exclude": ["swim", "pool", "aqua", "skatepark", "golf", "bowling",
                    "tennis center", "library", "marina"],
        "facility_label": "Baseball Field",
        "label_variants": {
            "softball": "Softball Field",
            "both": "Baseball/Softball Field",
        },
        "section_keywords": ["park", "field", "diamond"],
        "min_pitch_length_ft": 200,
    },
    "Basketball": {
        "osm_sports": ["basketball"],
        "keywords": ["basketball", "gym", "recreation center", "rec center",
                     "community center", "boys & girls", "boys and girls",
                     "sports centre", "ymca"],
        "exclude": ["swim", "pool", "aqua", "skatepark", "golf", "bowling",
                    "marina", "library", "model airplane"],
        "facility_label": "Basketball Court",
        "label_variants": {
            "gym": "Gymnasium Basketball Court",
            "half": "Half Court",
            "full": "Full Court",
        },
        "section_keywords": ["park", "court", "gym", "recreation"],
        "min_pitch_length_ft": 42,
    },
    "Tennis": {
        "osm_sports": ["tennis"],
        "keywords": ["tennis", "racquet", "racket club"],
        "exclude": ["swim", "pool", "aqua", "skatepark", "golf", "bowling",
                    "marina", "library"],
        "facility_label": "Tennis Court",
        "label_variants": {},
        "section_keywords": ["park", "court", "tennis", "club"],
        "min_pitch_length_ft": 75,
    },
    "Volleyball": {
        "osm_sports": ["volleyball", "beachvolleyball"],
        "keywords": ["volleyball", "beach volleyball"],
        "exclude": ["swim", "pool", "aqua", "skatepark", "golf", "bowling",
                    "marina", "library"],
        "facility_label": "Volleyball Court",
        "label_variants": {
            "beach": "Beach Volleyball Court",
        },
        "section_keywords": ["park", "court", "beach", "gym"],
        "min_pitch_length_ft": 55,
    },
}

# Per (sport, category) dimensions: (L1, W1, L2, W2). From GamePlay memo.
SPORT_DIMENSIONS = {
    "Baseball / Softball": {
        "ELEMENTARY SCHOOLS":            (210, 135, 225, 275),
        "MIDDLE SCHOOLS":                (210, 135, 300, 360),
        "HIGH SCHOOLS":                  (210, 135, 330, 390),
        "PUBLIC PARKS & RECREATION":     (210, 135, 300, 360),
        "COLLEGE":                       (210, 135, 225, 275),
        "GYMNASIUM / INDOOR FACILITIES": (210, 135, 225, 275),
        "OTHER FACILITIES":              (210, 135, 225, 275),
    },
    "Basketball": {
        "ELEMENTARY SCHOOLS":            (74, 50, None, None),
        "MIDDLE SCHOOLS":                (94, 50, None, None),
        "HIGH SCHOOLS":                  (94, 50, None, None),
        "PUBLIC PARKS & RECREATION":     (94, 50, None, None),
        "COLLEGE":                       (94, 50, None, None),
        "GYMNASIUM / INDOOR FACILITIES": (94, 50, None, None),
        "OTHER FACILITIES":              (94, 50, None, None),
    },
    "Soccer / Football": {
        "ELEMENTARY SCHOOLS":            (210, 135, 225, 275),
        "MIDDLE SCHOOLS":                (300, 150, 225, 275),
        "HIGH SCHOOLS":                  (300, 150, 225, 275),
        "PUBLIC PARKS & RECREATION":     (300, 150, 225, 275),
        "COLLEGE":                       (300, 150, 225, 275),
        "GYMNASIUM / INDOOR FACILITIES": (300, 150, 225, 275),
        "OTHER FACILITIES":              (300, 150, 225, 275),
    },
    "Tennis": {
        "ELEMENTARY SCHOOLS":            (78, 36, None, None),
        "MIDDLE SCHOOLS":                (78, 36, None, None),
        "HIGH SCHOOLS":                  (78, 36, None, None),
        "PUBLIC PARKS & RECREATION":     (78, 36, None, None),
        "COLLEGE":                       (78, 36, None, None),
        "GYMNASIUM / INDOOR FACILITIES": (78, 36, None, None),
        "OTHER FACILITIES":              (78, 36, None, None),
    },
    "Volleyball": {
        "ELEMENTARY SCHOOLS":            (50, 25, None, None),
        "MIDDLE SCHOOLS":                (60, 30, None, None),
        "HIGH SCHOOLS":                  (60, 30, None, None),
        "PUBLIC PARKS & RECREATION":     (60, 30, None, None),
        "COLLEGE":                       (60, 30, None, None),
        "GYMNASIUM / INDOOR FACILITIES": (60, 30, None, None),
        "OTHER FACILITIES":              (60, 30, None, None),
    },
}

SPORT_AGE_GROUP = {
    "Baseball / Softball": {
        "ELEMENTARY SCHOOLS": "12U", "MIDDLE SCHOOLS": "14U",
        "HIGH SCHOOLS": "18U", "PUBLIC PARKS & RECREATION": "14U",
        "COLLEGE": "12U", "GYMNASIUM / INDOOR FACILITIES": "12U",
        "OTHER FACILITIES": "12U",
    },
    "Basketball": {
        "ELEMENTARY SCHOOLS": "12U", "MIDDLE SCHOOLS": "18U",
        "HIGH SCHOOLS": "18U", "PUBLIC PARKS & RECREATION": "18U",
        "COLLEGE": "18U", "GYMNASIUM / INDOOR FACILITIES": "18U",
        "OTHER FACILITIES": "18U",
    },
    "Soccer / Football": {
        "ELEMENTARY SCHOOLS": "12U", "MIDDLE SCHOOLS": "18U",
        "HIGH SCHOOLS": "18U", "PUBLIC PARKS & RECREATION": "18U",
        "COLLEGE": "18U", "GYMNASIUM / INDOOR FACILITIES": "18U",
        "OTHER FACILITIES": "18U",
    },
    "Tennis": {k: "18U" for k in [
        "ELEMENTARY SCHOOLS","MIDDLE SCHOOLS","HIGH SCHOOLS",
        "PUBLIC PARKS & RECREATION","COLLEGE",
        "GYMNASIUM / INDOOR FACILITIES","OTHER FACILITIES"]},
    "Volleyball": {k: "18U" for k in [
        "ELEMENTARY SCHOOLS","MIDDLE SCHOOLS","HIGH SCHOOLS",
        "PUBLIC PARKS & RECREATION","COLLEGE",
        "GYMNASIUM / INDOOR FACILITIES","OTHER FACILITIES"]},
}

OTHER_PRIMARY_BY = {
    "Baseball / Softball": {"*": "Softball, 18U"},
    "Basketball":          {"*": ""},
    "Soccer / Football":   {"*": "Football 18U"},
    # "Soccer / Football": {
    #     "ELEMENTARY SCHOOLS": "Field hockey 12U, Lacrosse 12U, Rugby 12U, Ultimate Frisbee 12U",
    #     "*":                  "Field hockey 18U, Lacrosse 18U, Rugby 18U, Ultimate Frisbee 18U",
    # },
    "Tennis":     {"*": "Pickleball 18U"},
    "Volleyball": {"*": ""},
}

SECONDARY_BY = {
    "Baseball / Softball": {"*": "Soccer, 12U; Rugby, 12U; Football 12U, Field Hockey 12U, Ultimate 12U, Lacrosse 12U"},
    "Basketball":          {"*": ""},
    "Soccer / Football": {
        "ELEMENTARY SCHOOLS": "Field hockey 12U, Lacrosse 12U, Rugby 12U, Ultimate Frisbee 12U",
        "*":                  "Field hockey 18U, Lacrosse 18U, Rugby 18U, Ultimate Frisbee 18U",
    },
    # "Soccer / Football":   {"*": "Football 18U"},
    "Tennis":              {"*": ""},
    "Volleyball":          {"*": ""},
}

_TOO_SMALL_NAME_FRAGMENTS = [
    "tot lot", "tot-lot", "toddler", "mini park", "mini-park",
    "pocket park", "dog park", "dog run", "skate park", "skatepark",
    "splash pad", "spray park", "butterfly garden", "community garden",
    "meditation garden", "memorial garden", "rose garden",
]

def haversine(lat1, lon1, lat2, lon2):
    R = 6371000
    p = math.pi / 180
    a = (math.sin((lat2 - lat1) * p / 2) ** 2 +
         math.cos(lat1 * p) * math.cos(lat2 * p) *
         math.sin((lon2 - lon1) * p / 2) ** 2)
    return 2 * R * math.asin(math.sqrt(a))

def clean_name(name):
    if not name:
        return ""
    return re.sub(r"\s+", " ", name).strip()

def normalize_key(name):
    n = name.lower().strip()
    for suffix in [" park", " field", " fields", " court", " courts"]:
        if n.endswith(suffix):
            n = n[:-len(suffix)].strip()
    return re.sub(r"[^a-z0-9]", "", n)

NOMINATIM_MIRRORS = [
    "https://nominatim.openstreetmap.org",
]

def _photon_request(path, params, timeout=20, max_attempts=4):
    """Call Photon (Komoot) geocoder. Less throttled than Nominatim."""
    base = "https://photon.komoot.io"
    last_error = None
    for attempt in range(max_attempts):
        resp = None
        try:
            resp = requests.get(f"{base}{path}", params=params,
                                headers=HEADERS, timeout=timeout)
            if resp.status_code in (429, 503):
                last_error = f"Photon HTTP {resp.status_code}"
                time.sleep(2 ** (attempt + 1))
                continue
            resp.raise_for_status()
            return resp.json(), base
        except requests.exceptions.Timeout:
            last_error = f"Photon timeout (attempt {attempt+1})"
        except requests.exceptions.ConnectionError:
            last_error = f"Photon connection refused (attempt {attempt+1})"
        except requests.exceptions.HTTPError:
            status = resp.status_code if resp is not None else 0
            last_error = f"Photon HTTP {status}"
            break
        except Exception as e:
            last_error = f"Photon {type(e).__name__}: {e}"
            break
        if attempt < max_attempts - 1:
            time.sleep(2 ** (attempt + 1))
    raise RuntimeError(f"Photon unreachable. Last error: {last_error}")

_nominatim_request = _photon_request

def lookup_city_bbox(city, county, state="California", country="USA", use_cache=True):
    """Resolve city via Photon. Returns bbox + OSM relation/way id for
    Qlever's spatial join (`ogc:sfContains`)."""
    key = _cache_key("city_bbox_qlever_v2", city, county, state, country)
    if use_cache:
        cached = cache_get(key)
        if cached is not None:
            return cached

    target_lower = city.lower().strip()
    target_state_lower = (state or "").lower().strip()

    params = [("q", city), ("limit", "10"),
              ("osm_tag", "place:city"), ("osm_tag", "place:town"),
              ("osm_tag", "place:village"), ("osm_tag", "boundary:administrative")]
    try:
        data, _ = _photon_request("/api/", params)
    except RuntimeError as e:
        st.error(f"Photon lookup failed: {e}")
        return None

    valid_item = None
    for feat in data.get("features", []):
        props = feat.get("properties", {})
        name = (props.get("name") or "").lower()
        st_name = (props.get("state") or "").lower()
        if target_lower not in name and name not in target_lower:
            continue
        if target_state_lower and target_state_lower not in st_name:
            continue
        valid_item = feat
        break

    if not valid_item:
        return None

    props = valid_item.get("properties", {})
    geom = valid_item.get("geometry", {})
    coords = geom.get("coordinates", [])
    ext = props.get("extent")  # [west, north, east, south]
    if ext and len(ext) == 4:
        min_lon, max_lat, max_lon, min_lat = ext
    elif coords and len(coords) == 2:
        lon, lat = coords
        min_lat, max_lat = lat - 0.05, lat + 0.05
        min_lon, max_lon = lon - 0.05, lon + 0.05
    else:
        return None

    lat_span = max_lat - min_lat
    lon_span = max_lon - min_lon
    lat_buf = max(lat_span * 0.08, 0.005)
    lon_buf = max(lon_span * 0.08, 0.005)

    result = {
        "min_lat": min_lat - lat_buf,
        "max_lat": max_lat + lat_buf,
        "min_lon": min_lon - lon_buf,
        "max_lon": max_lon + lon_buf,
        "match_display": props.get("name", ""),
        "osm_type": props.get("osm_type", ""),   # "R" | "W" | "N"
        "osm_id": props.get("osm_id", 0),
    }
    if use_cache:
        cache_set(key, result)
    return result

def point_in_polygon(lat, lon, polygon):
    if not polygon or len(polygon) < 3:
        return True
    inside = False
    n = len(polygon)
    j = n - 1
    for i in range(n):
        lat_i, lon_i = polygon[i]
        lat_j, lon_j = polygon[j]
        if ((lon_i > lon) != (lon_j > lon)) and \
           (lat < (lat_j - lat_i) * (lon - lon_i) / (lon_j - lon_i + 1e-12) + lat_i):
            inside = not inside
        j = i
    return inside

class OverpassCircuitBreaker:
    def __init__(self):
        self._lock = Lock()
        self._tripped = False
        self._reason = ""

    def is_tripped(self):
        with self._lock:
            return self._tripped

    def trip(self, reason):
        with self._lock:
            if not self._tripped:
                self._tripped = True
                self._reason = reason

    def reason(self):
        with self._lock:
            return self._reason

    def reset(self):
        with self._lock:
            self._tripped = False
            self._reason = ""

QLEVER_ENDPOINT = "https://qlever.cs.uni-freiburg.de/api/osm-planet"

# Qlever exposes pre-computed spatial relations via ogc:sfContains. The free
# `geof:sfIntersects` function only works for a restricted subset of inputs
# on Qlever, so we join via the city's OSM relation/way id instead.
SPARQL_PREFIXES = """
PREFIX geo: <http://www.opengis.net/ont/geosparql#>
PREFIX ogc: <http://www.opengis.net/rdf#>
PREFIX osmkey: <https://www.openstreetmap.org/wiki/Key:>
PREFIX osm: <https://www.openstreetmap.org/>
PREFIX osmrel: <https://www.openstreetmap.org/relation/>
PREFIX osmway: <https://www.openstreetmap.org/way/>
PREFIX osmnode: <https://www.openstreetmap.org/node/>
PREFIX rdf: <http://www.w3.org/1999/02/22-rdf-syntax-ns#>
"""

_OSM_TYPE_PREFIX = {"R": "osmrel:", "W": "osmway:", "N": "osmnode:"}

def _city_uri(bbox):
    """Build the SPARQL term referring to the city as an OSM entity."""
    t = bbox.get("osm_type") or "R"
    oid = bbox.get("osm_id")
    if not oid:
        return None
    return f"{_OSM_TYPE_PREFIX.get(t, 'osmrel:')}{oid}"

_INDOOR_SPORTS = {"Basketball", "Volleyball"}

def build_qlever_queries(bbox, sport_config, sport_choice=""):
    """SPARQL queries with sport-aware container filtering.

    For each container type (parks, schools, sports_centres) we issue a
    pitch-containment query: `?facility ogc:sfContains ?pitch` where the
    pitch matches this sport. This returns only facilities that actually
    host the sport (per OSM data) — replacing the prior haversine 200m
    heuristic.

    For indoor sports we ALSO include all sports_centres/community_centres
    regardless of pitch tagging, because indoor courts are rarely tagged
    with a sport in OSM.
    """
    city = _city_uri(bbox)
    sports_filter = " || ".join(
        [f'CONTAINS(LCASE(STR(?sport)), "{s.lower()}")'
         for s in sport_config["osm_sports"]]
    )
    spatial = f"{city} ogc:sfContains ?osm_id ." if city else ""

    pitches = f"""{SPARQL_PREFIXES}
SELECT ?osm_id ?name ?sport ?lit ?hoops ?wkt WHERE {{
  {spatial}
  ?osm_id osmkey:leisure "pitch" ;
          osmkey:sport ?sport ;
          geo:hasGeometry/geo:asWKT ?wkt .
  OPTIONAL {{ ?osm_id osmkey:name ?name }}
  OPTIONAL {{ ?osm_id osmkey:lit ?lit }}
  OPTIONAL {{ ?osm_id osmkey:hoops ?hoops }}
  FILTER({sports_filter})
}}
LIMIT 5000"""

    # v5: drop the `ogc:sfContains ?pitch` requirement. Fetch ALL parks /
    # schools / sports_centres / recreation_grounds in the city bbox; the
    # post-merge proximity step attaches any nearby pitch (500m radius).
    # Schools etc with no OSM-tagged pitch nearby get filtered out by the
    # `has_pitches or is_sport_name` check in merge_and_deduplicate.
    # This restores middle schools and other facilities that exist in OSM
    # but lack the explicit leisure=pitch child geometry.
    parks = f"""{SPARQL_PREFIXES}
SELECT DISTINCT ?osm_id ?name ?wkt WHERE {{
  {spatial}
  ?osm_id osmkey:leisure "park" ;
          osmkey:name ?name ;
          geo:hasGeometry/geo:asWKT ?wkt .
}}
LIMIT 2000"""

    schools = f"""{SPARQL_PREFIXES}
SELECT DISTINCT ?osm_id ?name ?amenity ?wkt WHERE {{
  {spatial}
  ?osm_id osmkey:amenity ?amenity ;
          osmkey:name ?name ;
          geo:hasGeometry/geo:asWKT ?wkt .
  FILTER(?amenity IN ("school", "college", "university"))
}}
LIMIT 2000"""

    sports_centres = f"""{SPARQL_PREFIXES}
SELECT DISTINCT ?osm_id ?name ?leisure ?wkt WHERE {{
  {spatial}
  ?osm_id osmkey:leisure ?leisure ;
          osmkey:name ?name ;
          geo:hasGeometry/geo:asWKT ?wkt .
  FILTER(?leisure IN ("sports_centre", "fitness_centre", "stadium"))
}}
LIMIT 1000"""

    rec_grounds = f"""{SPARQL_PREFIXES}
SELECT DISTINCT ?osm_id ?name ?wkt WHERE {{
  {spatial}
  ?osm_id osmkey:landuse "recreation_ground" ;
          osmkey:name ?name ;
          geo:hasGeometry/geo:asWKT ?wkt .
}}
LIMIT 1000"""

    queries = {
        f"{sport_config['facility_label']}s": ("pitch", pitches),
        "Parks": ("park", parks),
        "Schools": ("school", schools),
        "Sports centres + stadiums": ("sports_centre", sports_centres),
        "Recreation grounds": ("park", rec_grounds),
    }

    # Indoor sports: include all sports_centres/community_centres/sports_halls
    # even without OSM pitch tags (indoor courts rarely tagged).
    if sport_choice in _INDOOR_SPORTS:
        indoor_query = f"""{SPARQL_PREFIXES}
SELECT DISTINCT ?osm_id ?name ?leisure ?amenity ?building ?wkt WHERE {{
  {spatial}
  ?osm_id osmkey:name ?name ;
          geo:hasGeometry/geo:asWKT ?wkt .
  OPTIONAL {{ ?osm_id osmkey:leisure ?leisure }}
  OPTIONAL {{ ?osm_id osmkey:amenity ?amenity }}
  OPTIONAL {{ ?osm_id osmkey:building ?building }}
  FILTER(
    ?leisure IN ("sports_centre", "fitness_centre", "sports_hall") ||
    ?amenity IN ("community_centre", "gym") ||
    ?building IN ("sports_hall", "gymnasium")
  )
}}
LIMIT 1500"""
        queries["Indoor gyms / community centres"] = ("sports_centre", indoor_query)

    return queries

def _parse_wkt_point(wkt):
    """Return (lat, lon) from a WKT geometry. Centroid for polygons."""
    if not wkt:
        return None, None
    s = wkt.strip()
    if s.startswith("POINT"):
        m = re.search(r"POINT\s*\(\s*(-?\d+\.?\d*)\s+(-?\d+\.?\d*)", s)
        if m:
            return float(m.group(2)), float(m.group(1))
    nums = re.findall(r"-?\d+\.\d+", s)
    if len(nums) >= 2:
        lons = [float(x) for x in nums[0::2]]
        lats = [float(y) for y in nums[1::2]]
        return sum(lats)/len(lats), sum(lons)/len(lons)
    return None, None

def query_qlever(name, query, status_callback, use_cache=True, timeout=90):
    """POST SPARQL to Qlever. Returns (name, list-of-row-dicts)."""
    if use_cache:
        k = _cache_key("qlever", QLEVER_ENDPOINT, query)
        cached = cache_get(k)
        if cached is not None:
            with _log_lock:
                status_callback(f"  [{name}] cached ({len(cached)} rows)")
            return name, cached

    headers = {**HEADERS, "Accept": "application/sparql-results+json"}
    try:
        with _log_lock:
            status_callback(f"  [{name}] querying Qlever...")
        # Global semaphore: max 3 concurrent SPARQL POSTs to public Qlever.
        with _QLEVER_SEM:
            resp = requests.post(QLEVER_ENDPOINT, data={"query": query},
                                  headers=headers, timeout=timeout)
        resp.raise_for_status()
        payload = resp.json()
        rows = payload.get("results", {}).get("bindings", [])
        flat = []
        for r in rows:
            flat.append({k: v.get("value", "") for k, v in r.items()})
        with _log_lock:
            status_callback(f"  [{name}] OK {len(flat)} rows")
        # Only cache non-empty results. Empty rows often indicate transient
        # Qlever throttle; caching them poisons the next 7 days of runs.
        if use_cache and flat:
            cache_set(_cache_key("qlever", QLEVER_ENDPOINT, query), flat)
        return name, flat
    except Exception as e:
        with _log_lock:
            status_callback(f"  [{name}] ERR {type(e).__name__}: {e}")
        return name, []

def _resolve_sport_choice(sport_config):
    """Recover the SPORTS_CONFIG key from a sport_config dict."""
    for k, v in SPORTS_CONFIG.items():
        if v is sport_config:
            return k
    return ""

def fetch_overpass(bbox, sport_config, overpass_url, status_callback,
                    is_local=False, use_cache=True):
    """Fetch facilities via Qlever SPARQL (Overpass replacement)."""
    status_callback("Source 1: Qlever SPARQL (OSM-planet)...")
    sport_choice = _resolve_sport_choice(sport_config)
    queries = build_qlever_queries(bbox, sport_config, sport_choice)
    results = []
    with ThreadPoolExecutor(max_workers=3) as ex:
        futures = {
            ex.submit(query_qlever, qname, qtext, status_callback, use_cache):
                (qname, kind)
            for qname, (kind, qtext) in queries.items()
        }
        for fut in as_completed(futures):
            qname, kind = futures[fut]
            try:
                _, rows = fut.result()
                for r in rows:
                    lat, lon = _parse_wkt_point(r.get("wkt", ""))
                    if lat is None:
                        continue
                    leisure = "pitch" if kind == "pitch" else (
                        "park" if kind == "park" else (
                            r.get("leisure", "") if kind == "sports_centre" else ""
                        )
                    )
                    amenity = r.get("amenity", "") if kind == "school" else ""
                    results.append({
                        "source": "qlever",
                        "name": clean_name(r.get("name", "")),
                        "lat": lat, "lon": lon,
                        "sport": r.get("sport", ""),
                        "leisure": leisure,
                        "amenity": amenity,
                        "building": "",
                        "tags": {
                            "lit": r.get("lit", ""),
                            "hoops": r.get("hoops", ""),
                        },
                        "length_ft": None,
                        "width_ft": None,
                    })
            except Exception as e:
                status_callback(f"  Worker error: {e}")

    status_callback(f"  Qlever total: {len(results)} raw elements")
    return results

def probe_overpass(overpass_url, timeout=5):
    return overpass_url, True, "qlever-mode"

def pick_working_overpass(preferred_url, status_callback, is_local=False):
    return QLEVER_ENDPOINT

def build_nominatim_searches(city, state, sport_config):
    facility = sport_config["facility_label"].lower()
    return [
        f"{facility} {city} {state}",
        f"{facility}s {city} {state}",
        f"{sport_config['osm_sports'][0]} {city} {state}",
        f"sports field {city} {state}",
        f"park {city} {state}",
        f"high school {city} {state}",
        f"middle school {city} {state}",
        f"elementary school {city} {state}",
        f"college {city} {state}",
        f"recreation center {city} {state}",
        f"community center {city} {state}",
        f"sports complex {city} {state}",
        f"playground {city} {state}",
    ]

def fetch_nominatim(city, state, bbox, sport_config, status_callback, use_cache=True):
    """No-op in Qlever mode — Qlever already returns named facilities."""
    status_callback("Source 2: Nominatim — skipped (Qlever provides names).")
    return []

NCES_API_BASE = "https://educationdata.urban.org/api/v1/schools/ccd/directory"
NCES_YEAR = 2020
NCES_CACHE_TTL_SECONDS = 30 * 24 * 3600

_US_STATE_ABBR = {
    "alabama":"AL","alaska":"AK","arizona":"AZ","arkansas":"AR",
    "california":"CA","colorado":"CO","connecticut":"CT","delaware":"DE",
    "florida":"FL","georgia":"GA","hawaii":"HI","idaho":"ID",
    "illinois":"IL","indiana":"IN","iowa":"IA","kansas":"KS",
    "kentucky":"KY","louisiana":"LA","maine":"ME","maryland":"MD",
    "massachusetts":"MA","michigan":"MI","minnesota":"MN","mississippi":"MS",
    "missouri":"MO","montana":"MT","nebraska":"NE","nevada":"NV",
    "new hampshire":"NH","new jersey":"NJ","new mexico":"NM","new york":"NY",
    "north carolina":"NC","north dakota":"ND","ohio":"OH","oklahoma":"OK",
    "oregon":"OR","pennsylvania":"PA","rhode island":"RI",
    "south carolina":"SC","south dakota":"SD","tennessee":"TN","texas":"TX",
    "utah":"UT","vermont":"VT","virginia":"VA","washington":"WA",
    "west virginia":"WV","wisconsin":"WI","wyoming":"WY",
    "district of columbia":"DC",
}

def _state_to_abbr(state):
    s = (state or "").strip()
    if len(s) == 2:
        return s.upper()
    return _US_STATE_ABBR.get(s.lower(), s[:2].upper())

def _fetch_nces_state_raw(state_abbr, use_cache=True):
    """Fetch every public school in a state (paginated). 30-day cache."""
    key = _cache_key("nces_state_v1", NCES_YEAR, state_abbr)
    if use_cache:
        cached = cache_get(key, max_age_seconds=NCES_CACHE_TTL_SECONDS)
        if cached is not None:
            return cached

    all_results = []
    url = f"{NCES_API_BASE}/{NCES_YEAR}/"
    params = {"state_location": state_abbr}
    for _page in range(20):
        try:
            resp = requests.get(url, params=params, headers=HEADERS, timeout=90)
            resp.raise_for_status()
            data = resp.json()
        except Exception:
            break
        all_results.extend(data.get("results", []))
        nxt = data.get("next")
        if not nxt:
            break
        url = nxt
        params = None

    if use_cache and all_results:
        cache_set(_cache_key("nces_state_v1", NCES_YEAR, state_abbr), all_results)
    return all_results

def _nces_level_to_category(lvl):
    try:
        lvl = int(lvl)
    except (TypeError, ValueError):
        return "OTHER FACILITIES"
    return {
        1: "ELEMENTARY SCHOOLS",
        2: "MIDDLE SCHOOLS",
        3: "HIGH SCHOOLS",
        4: "OTHER FACILITIES",
    }.get(lvl, "OTHER FACILITIES")

def fetch_nces_schools(target_city, target_state, bbox, status_callback,
                        use_cache=True):
    """Fetch NCES public schools in target_city; returns entries in the
    same schema Qlever produces so downstream merge/geocode/Excel is
    unchanged. Every school is auto-categorized via school_level."""
    state_abbr = _state_to_abbr(target_state)
    status_callback(f"Source 2: NCES CCD schools ({state_abbr})...")
    all_rows = _fetch_nces_state_raw(state_abbr, use_cache=use_cache)
    if not all_rows:
        status_callback(f"  NCES: 0 rows for {state_abbr}")
        return []

    target = (target_city or "").strip().lower()
    matched = []
    for r in all_rows:
        loc = (r.get("city_location") or "").strip().lower()
        lat = r.get("latitude")
        lon = r.get("longitude")
        try:
            lat = float(lat) if lat is not None else None
            lon = float(lon) if lon is not None else None
        except (TypeError, ValueError):
            lat = lon = None
        if lat is None or lon is None:
            continue
        # Match by city string OR bbox — handles adjacent cities & aliases.
        in_bbox_ok = (bbox and
                      bbox["min_lat"] <= lat <= bbox["max_lat"] and
                      bbox["min_lon"] <= lon <= bbox["max_lon"])
        if loc != target and not in_bbox_ok:
            continue

        name = (r.get("school_name") or "").strip()
        if not name:
            continue
        matched.append({
            "source": "nces",
            "name": clean_name(name),
            "lat": lat, "lon": lon,
            "sport": "",
            "leisure": "",
            "amenity": "school",
            "building": "",
            "tags": {
                "addr:city": r.get("city_location", ""),
                "nces_level": r.get("school_level", ""),
            },
            "nces_level": r.get("school_level", ""),
            "length_ft": None,
            "width_ft": None,
        })
    status_callback(f"  NCES matched: {len(matched)} schools in {target_city}")
    return matched

def is_confirmed_sport(entry, sport_config):
    sport = entry.get("sport", "").lower()
    return any(s in sport for s in sport_config["osm_sports"])

def is_facility(entry):
    leisure = entry.get("leisure", "").lower()
    amenity = entry.get("amenity", "").lower()
    building = entry.get("building", "").lower()
    name = entry.get("name", "").lower()
    return (leisure in ("park", "sports_centre", "fitness_centre",
                        "sports_hall", "stadium", "school", "college",
                        "university", "recreation_ground", "playground") or
            amenity in ("school", "college", "university",
                        "community_centre", "leisure", "amenity") or
            building in ("sports_hall",) or
            entry.get("source") == "nominatim" or
            any(k in name for k in ["park", "school", "college",
                                     "recreation", "field", "playground"]))

def is_too_small(entry, sport_config):
    name = entry.get("name", "").lower()
    if any(frag in name for frag in _TOO_SMALL_NAME_FRAGMENTS):
        return True

    leisure = entry.get("leisure", "").lower()
    is_confirmed = is_confirmed_sport(entry, sport_config)
    if leisure == "playground" and not is_confirmed:
        return True

    length_ft = entry.get("length_ft")
    if length_ft:
        min_ft = sport_config.get("min_pitch_length_ft", 0)
        if length_ft < min_ft:
            return True

    children = entry.get("child_pitches", [])
    if children:
        min_ft = sport_config.get("min_pitch_length_ft", 0)
        all_small = all(
            (c.get("length_ft") or 0) > 0 and (c.get("length_ft") or 0) < min_ft
            for c in children
        )
        if all_small:
            return True

    return False

def merge_and_deduplicate(all_sources, sport_config, status_callback):
    status_callback(f"Total raw entries: {len(all_sources)}")

    coord_seen = set()
    deduped = []
    for entry in all_sources:
        if entry.get("lat") and entry.get("lon"):
            ck = (round(entry["lat"], 5), round(entry["lon"], 5))
            if ck in coord_seen:
                continue
            coord_seen.add(ck)
        deduped.append(entry)
    status_callback(f"After coord dedup: {len(deduped)}")

    confirmed = []
    facilities = []
    for entry in deduped:
        # Pitch/court rows come from the sport-filtered SPARQL query. Keep
        # only those whose sport tag confirms this sport; drop the rest.
        if entry.get("leisure") == "pitch":
            if is_confirmed_sport(entry, sport_config):
                confirmed.append(entry)
            continue
        if is_facility(entry):
            facilities.append(entry)

    status_callback(f"Confirmed pitches: {len(confirmed)}")
    status_callback(f"Facilities: {len(facilities)}")

    fac_seen = {}
    exclude_list = sport_config["exclude"]
    for entry in facilities:
        name = entry["name"].strip()
        if not name:
            continue
        if any(k in name.lower() for k in exclude_list):
            continue
        key = normalize_key(name)
        if not key:
            continue
        if key not in fac_seen:
            fac_seen[key] = entry
        else:
            existing = fac_seen[key]
            if not existing.get("lat") and entry.get("lat"):
                entry["name"] = entry["name"] or existing["name"]
                fac_seen[key] = entry

    facility_list = list(fac_seen.values())

    _SCHOOL_PRIORITY = {
        "high school": 0, "high sch": 0, "preparatory": 0, "prep school": 0,
        "middle school": 1, "middle sch": 1, "junior high": 1, "intermediate": 1,
        "elementary": 2, "primary school": 2,
        "college": 3, "university": 3,
    }

    def _school_priority(name):
        n = name.lower()
        for kw, rank in _SCHOOL_PRIORITY.items():
            if kw in n:
                return rank
        return 99

    def _institution_stem(name):
        n = name.lower()
        for kw in ["high school", "middle school", "junior high", "elementary school",
                   "elementary", "primary school", "preparatory", "prep school",
                   "college", "university", "intermediate school", "intermediate"]:
            n = n.replace(kw, "").strip()
        return re.sub(r"[^a-z0-9]", "", n)

    SAME_CAMPUS_RADIUS = 60
    suppressed = set()
    fl = facility_list
    for i in range(len(fl)):
        if i in suppressed:
            continue
        for j in range(i + 1, len(fl)):
            if j in suppressed:
                continue
            a, b = fl[i], fl[j]
            if not (a.get("lat") and a.get("lon") and b.get("lat") and b.get("lon")):
                continue
            dist = haversine(a["lat"], a["lon"], b["lat"], b["lon"])
            if dist > SAME_CAMPUS_RADIUS:
                continue
            stem_a = _institution_stem(a.get("name", ""))
            stem_b = _institution_stem(b.get("name", ""))
            if not stem_a or not stem_b or stem_a != stem_b:
                continue
            pri_a = _school_priority(a.get("name", ""))
            pri_b = _school_priority(b.get("name", ""))
            if pri_a == 99 and pri_b == 99:
                continue
            if pri_a <= pri_b:
                suppressed.add(j)
            else:
                suppressed.add(i)

    if suppressed:
        status_callback(f"  Same-campus dedup removed {len(suppressed)} co-located duplicate(s)")
        facility_list = [fl[i] for i in range(len(fl)) if i not in suppressed]

    # 500m radius covers large high-school + college campuses. The previous
    # 200m left edge-of-campus pitches stranded as standalone entries
    # (e.g. soccer pitch on Westmoor High School's far field).
    PROXIMITY_RADIUS = 500
    for pitch in confirmed:
        if not pitch.get("lat") or not pitch.get("lon"):
            continue
        best_fac = None
        best_dist = PROXIMITY_RADIUS + 1
        for fac in facility_list:
            if not fac.get("lat") or not fac.get("lon"):
                continue
            dist = haversine(pitch["lat"], pitch["lon"], fac["lat"], fac["lon"])
            if dist < best_dist:
                best_dist = dist
                best_fac = fac
        if best_fac:
            if "child_pitches" not in best_fac:
                best_fac["child_pitches"] = []
            best_fac["child_pitches"].append(pitch)
        else:
            name = pitch.get("name", "")
            if name:
                key = normalize_key(name)
                if key and key not in fac_seen:
                    fac_seen[key] = pitch
                    facility_list.append(pitch)
            elif pitch.get("lat"):
                pitch["name"] = f"{sport_config['facility_label']} ({pitch['lat']:.4f}, {pitch['lon']:.4f})"
                facility_list.append(pitch)

    # v5 fix: the SPARQL fetch no longer requires `ogc:sfContains ?pitch`,
    # so facility_list contains every park/school in the bbox regardless
    # of sport. Restore the sport-gating post-filter: keep a facility only
    # if it has an attached pitch of this sport (via the 500m proximity
    # step above) or its name explicitly matches the sport's keywords.
    # Without this, the tennis sheet listed every park in the city
    # (v5 hallucination bug).
    results = []
    for fac in facility_list:
        has_pitches = len(fac.get("child_pitches", [])) > 0
        name_lower = fac.get("name", "").lower()
        is_sport_name = any(k in name_lower for k in sport_config["keywords"])
        # NCES-sourced schools bypass the sport-gating filter: every real
        # public school is a valid facility for the memo sports and gets
        # default dims from SPORT_DIMENSIONS even without OSM pitch data.
        is_nces = fac.get("source") == "nces"
        if has_pitches or is_sport_name or is_nces:
            results.append(fac)

    results = [r for r in results if r.get("name")]

    before_size = len(results)
    results = [r for r in results if not is_too_small(r, sport_config)]
    removed_small = before_size - len(results)
    if removed_small:
        status_callback(f"  Removed {removed_small} too-small / wrong-type entries "
                        f"(tot lots, playgrounds, undersized pitches)")

    multi = sum(1 for r in results if len(r.get("child_pitches", [])) > 1)
    status_callback(f"After merge: {len(results)} facilities ({multi} multi-court/field)")
    return results

def _fallback_address(target_city, target_state="", postcode="", lat=None, lon=None):
    state_abbr = ""
    if target_state:
        s = target_state.strip()
        state_abbr = s[:2].upper() if len(s) >= 2 else s.upper()
    parts = [p for p in [target_city, state_abbr, postcode] if p]
    out = ", ".join([parts[0]] + ([" ".join(parts[1:])] if len(parts) > 1 else []))
    if lat is not None and lon is not None:
        out = f"{out} (@ {lat:.4f}, {lon:.4f})".strip()
    return out

def _reverse_geocode_one(entry, target_city, nominatim_url=None,
                           use_cache=True, target_state="", retries=2):
    """Reverse geocode via Photon."""
    tags = entry.get("tags", {})
    street = tags.get("addr:street", "")
    number = tags.get("addr:housenumber", "")
    if street:
        city = tags.get("addr:city", target_city)
        state_tag = tags.get("addr:state", "")
        postcode_tag = tags.get("addr:postcode", "")
        state_abbr = state_tag[:2].upper() if state_tag else (
            target_state[:2].upper() if target_state else "")
        street_line = f"{number} {street}".strip()
        loc_line = f"{city}, {state_abbr} {postcode_tag}".strip().rstrip(",")
        entry["address"] = f"{street_line}, {loc_line}".strip(", ")
        entry["verified_city"] = city.lower()
        entry["zipcode"] = postcode_tag
        return entry, "osm_tags"

    lat_r = round(entry["lat"], 5)
    lon_r = round(entry["lon"], 5)
    if use_cache:
        key = _cache_key("photon_reverse_v1", lat_r, lon_r)
        cached = cache_get(key)
        if cached is not None:
            entry["address"] = cached.get("address") or _fallback_address(
                target_city, target_state, lat=entry["lat"], lon=entry["lon"])
            entry["verified_city"] = cached.get("verified_city", "")
            entry["zipcode"] = cached.get("zipcode", "")
            return entry, "cached"

    last_err = None
    for attempt in range(retries + 1):
        try:
            params = {"lat": entry["lat"], "lon": entry["lon"], "lang": "en"}
            data, _ = _photon_request("/reverse", params, timeout=20)
            feats = data.get("features", [])
            if not feats:
                raise RuntimeError("no features")
            props = feats[0].get("properties", {})
            road = props.get("street", "") or props.get("name", "")
            house = props.get("housenumber", "")
            city = props.get("city", "") or props.get("town", "") or props.get("village", "")
            postcode = props.get("postcode", "")
            state = props.get("state", "")
            entry["verified_city"] = city.lower() if city else ""
            entry["zipcode"] = postcode
            display_city = city if city else target_city
            state_abbr = state[:2].upper() if state else (
                target_state[:2].upper() if target_state else "")
            if road:
                street_line = f"{house} {road}".strip()
                loc_line = f"{display_city}, {state_abbr} {postcode}".strip().rstrip(",")
                entry["address"] = f"{street_line}, {loc_line}".strip(", ")
            else:
                entry["address"] = _fallback_address(
                    display_city, target_state, postcode,
                    lat=entry["lat"], lon=entry["lon"])
            if use_cache:
                cache_set(_cache_key("photon_reverse_v1", lat_r, lon_r), {
                    "address": entry["address"],
                    "verified_city": entry["verified_city"],
                    "zipcode": entry["zipcode"],
                })
            return entry, "api"
        except Exception as e:
            last_err = e
            if attempt < retries:
                time.sleep(0.8 * (attempt + 1))
            continue
    entry["address"] = _fallback_address(
        target_city, target_state, lat=entry["lat"], lon=entry["lon"])
    entry["verified_city"] = ""
    entry["zipcode"] = ""
    return entry, "api_failed"

def reverse_geocode_all(entries, target_city, status_callback,
                         use_local_nominatim=False, use_cache=True,
                         target_state=""):
    nominatim_url = NOMINATIM_URL
    n = len(entries)
    status_callback(f"Reverse geocoding {n} facilities...")

    osm_count = 0
    cache_count = 0
    api_needed = []
    for entry in entries:
        tags = entry.get("tags", {})
        if tags.get("addr:street"):
            _reverse_geocode_one(entry, target_city, nominatim_url, use_cache,
                                  target_state=target_state)
            osm_count += 1
            continue
        if use_cache:
            lat_r = round(entry["lat"], 5)
            lon_r = round(entry["lon"], 5)
            key = _cache_key("reverse_geocode_v2", lat_r, lon_r)
            cached = cache_get(key)
            if cached is not None:
                entry["address"] = cached.get("address") or _fallback_address(
                    target_city, target_state, lat=entry["lat"], lon=entry["lon"])
                entry["verified_city"] = cached.get("verified_city", "")
                entry["zipcode"] = cached.get("zipcode", "")
                cache_count += 1
                continue
        api_needed.append(entry)

    if osm_count:
        status_callback(f"  Used OSM addr tags: {osm_count}")
    if cache_count:
        status_callback(f"  Used cache: {cache_count} 💾")

    if not api_needed:
        status_callback(f"  All {n} addresses resolved without API calls")
        return

    if use_local_nominatim:
        status_callback(f"  Parallel reverse geocode {len(api_needed)} (local)...")
        with ThreadPoolExecutor(max_workers=8) as executor:
            list(executor.map(
                lambda e: _reverse_geocode_one(e, target_city, nominatim_url,
                                                use_cache, target_state=target_state),
                api_needed
            ))
    else:
        status_callback(f"  Sequential reverse geocode {len(api_needed)} "
                        f"(public Nominatim, 1 req/sec)...")
        for i, entry in enumerate(api_needed, 1):
            _reverse_geocode_one(entry, target_city, nominatim_url, use_cache,
                                  target_state=target_state)
            if i % 10 == 0:
                status_callback(f"    progress: {i}/{len(api_needed)}")
            time.sleep(1.1)

    status_callback(f"  Reverse geocoding complete")

def get_city_neighborhoods(city, state, country="USA", use_cache=True):
    """Simplified in Qlever/Photon mode — return just the city alias set."""
    return {city.lower()}

def _parse_city_from_address(address):
    if not address:
        return ""

    parts = [p.strip() for p in address.split(",")]
    if len(parts) < 2:
        return ""

    for part in parts[1:]:
        part = part.strip()
        if not part:
            continue
        if re.match(r"^\d{5}(-\d{4})?$", part):
            continue
        if re.match(r"^[A-Z]{2}(\s+\d{5}(-\d{4})?)?$", part):
            continue
        if re.match(r"^\d+$", part):
            continue
        if part.lower() in ("usa", "united states", "us"):
            continue
        city_token = re.split(r"\s+[A-Z]{2}\s+\d{5}", part)[0].strip()
        if city_token:
            return city_token.lower()

    return ""

def _parse_zip_from_address(address):
    if not address:
        return ""
    m = re.search(r"\b(\d{5}(?:-\d{4})?)\b", address)
    return m.group(1) if m else ""

def filter_wrong_city(entries, target_city, target_state, bbox, status_callback,
                       use_cache=True):
    status_callback(f"Filtering facilities outside {target_city}...")
    target = target_city.lower().strip()
    polygon = bbox.get("polygon") if bbox else None

    valid_aliases = get_city_neighborhoods(target_city, target_state,
                                            use_cache=use_cache)

    if polygon:
        status_callback(f"  Using city polygon ({len(polygon)} points) "
                        f"as secondary check")
    else:
        status_callback(f"  No polygon available — using address + verified_city")

    if len(valid_aliases) > 1:
        sample = ", ".join(sorted(valid_aliases))[:80]
        status_callback(f"  Recognizing {len(valid_aliases)} aliases: {sample}")

    filtered = []
    removed = []

    for entry in entries:
        lat = entry.get("lat")
        lon = entry.get("lon")
        v_city = entry.get("verified_city", "").lower().strip()
        address = entry.get("address", "")

        address_city = _parse_city_from_address(address)

        if address_city:
            if address_city in valid_aliases or target in address_city or address_city in target:
                filtered.append(entry)
                continue
            else:
                removed.append(
                    f"{entry['name']} (address city: '{address_city}' ≠ '{target}')"
                )
                continue

        if polygon and lat and lon:
            if point_in_polygon(lat, lon, polygon):
                filtered.append(entry)
                continue
            else:
                removed.append(f"{entry['name']} (outside city polygon)")
                continue

        if v_city:
            if v_city in valid_aliases or target in v_city or v_city in target:
                filtered.append(entry)
            else:
                removed.append(f"{entry['name']} (verified city: '{v_city}' ≠ '{target}')")
            continue

        filtered.append(entry)

    status_callback(f"Removed {len(removed)} facilities outside {target_city}")
    status_callback(f"Kept {len(filtered)} facilities")
    return filtered, removed

def categorize(entries, sport_config, status_callback):
    exclude = sport_config["exclude"]
    entries = [e for e in entries if not any(k in e["name"].lower() for k in exclude)]

    categories = {
        "PUBLIC PARKS & RECREATION": [],
        "GYMNASIUM / INDOOR FACILITIES": [],
        "HIGH SCHOOLS": [],
        "MIDDLE SCHOOLS": [],
        "ELEMENTARY SCHOOLS": [],
        "COLLEGE": [],
        "OTHER FACILITIES": [],
    }

    for entry in entries:
        # NCES entries carry an explicit school_level (1=Elem, 2=Mid, 3=Hi,
        # 4=Other) — use that first; falls back to name matching below.
        if entry.get("source") == "nces" and entry.get("nces_level") not in (None, ""):
            categories[_nces_level_to_category(entry.get("nces_level"))].append(entry)
            continue

        combined = (entry["name"] + " " + entry.get("address", "")).lower()
        if any(k in combined for k in ["high school", "high sch", "preparatory", "prep school"]):
            categories["HIGH SCHOOLS"].append(entry)
        elif any(k in combined for k in ["middle school", "middle sch", "junior high", "intermediate"]):
            categories["MIDDLE SCHOOLS"].append(entry)
        elif any(k in combined for k in ["elementary", "primary school"]):
            categories["ELEMENTARY SCHOOLS"].append(entry)
        elif any(k in combined for k in ["college", "university"]):
            categories["COLLEGE"].append(entry)
        elif any(k in combined for k in ["gym", "recreation center", "rec center",
                                          "community center", "boys & girls",
                                          "boys and girls", "sports centre",
                                          "sports center", "fitness", "indoor",
                                          "ymca"]):
            categories["GYMNASIUM / INDOOR FACILITIES"].append(entry)
        elif any(k in combined for k in ["park", "field", "memorial", "playground", "recreation"]):
            categories["PUBLIC PARKS & RECREATION"].append(entry)
        else:
            categories["OTHER FACILITIES"].append(entry)

    return categories

def _dims(sport_choice, category):
    return SPORT_DIMENSIONS.get(sport_choice, {}).get(category, (None, None, None, None))

def _lookup_by_category(mapping, sport_choice, category):
    table = mapping.get(sport_choice, {})
    if category in table:
        return table[category]
    return table.get("*", "")

def expand_to_rows(entries, sport_config, category_name, sport_choice):
    rows = []
    label = sport_config["facility_label"]
    variants = sport_config["label_variants"]
    L1, W1, L2, W2 = _dims(sport_choice, category_name)
    primary_sport = sport_choice.split(" / ")[0]
    other_primary = _lookup_by_category(OTHER_PRIMARY_BY, sport_choice, category_name)
    secondary_sport = _lookup_by_category(SECONDARY_BY, sport_choice, category_name)
    age_group = SPORT_AGE_GROUP.get(sport_choice, {}).get(category_name, "18U")

    for entry in entries:
        children = entry.get("child_pitches", [])
        num = max(len(children), 1)
        name_lower = entry["name"].lower()
        is_gym = (category_name == "GYMNASIUM / INDOOR FACILITIES" or
                  "gym" in name_lower)

        zipcode = entry.get("zipcode") or _parse_zip_from_address(entry.get("address", ""))

        def _desc(sport_str):
            d = label
            if "softball" in sport_str and "baseball" in sport_str and "both" in variants:
                d = variants["both"]
            elif "softball" in sport_str and "softball" in variants:
                d = variants["softball"]
            elif "football" in sport_str and "soccer" in sport_str and "soccer_football" in variants:
                d = variants["soccer_football"]
            elif "football" in sport_str and "soccer" not in sport_str and "football_only" in variants:
                d = variants["football_only"]
            elif is_gym and "gym" in variants:
                d = variants["gym"]
            elif "multi" in name_lower and "multi" in variants:
                d = variants["multi"]
            return d

        def _row(desc, lat, lon):
            return {
                "name": entry["name"],
                "description": desc,
                "address": entry.get("address", ""),
                "lat": lat,
                "lon": lon,
                "length_1_ft": L1,
                "width_1_ft":  W1,
                "length_2_ft": L2,
                "width_2_ft":  W2,
                "primary_sport": primary_sport,
                "age_group": age_group,
                "other_primary": other_primary,
                "secondary_sport": secondary_sport,
                "category": category_name,
                "zipcode": zipcode,
                "google_earth_url": f"https://earth.google.com/web/@{lat},{lon},50a,300d,35y,0h,0t,0r",
            }

        if num <= 1 and not children:
            sport = entry.get("sport", "").lower()
            desc = _desc(sport)
            tags = entry.get("tags", {})
            if tags.get("lit", "") == "yes":
                desc += " (Lighted)"
            rows.append(_row(desc, entry["lat"], entry["lon"]))
        else:
            for i, child in enumerate(children, 1):
                tags = child.get("tags", {})
                child_sport = child.get("sport", "").lower()
                base = _desc(child_sport)

                suffix_parts = []
                hoops = tags.get("hoops", "")
                if hoops == "1" and "half" in variants:
                    suffix_parts.append(variants["half"])
                elif hoops == "2" and "full" in variants:
                    suffix_parts.append(variants["full"])
                if tags.get("lit", "") == "yes":
                    suffix_parts.append("Lighted")

                desc = f"{base} {i}" if num > 1 else base
                if suffix_parts:
                    desc += f" ({', '.join(suffix_parts)})"

                rows.append(_row(
                    desc,
                    child.get("lat", entry["lat"]),
                    child.get("lon", entry["lon"]),
                ))

    # Excel-only augmentation: for Basketball at any non-outdoor category,
    # synthesize a Gymnasium row per facility (indoor counterpart to the
    # outdoor courts found via OSM). Volleyball 18U goes in Secondary
    # Sports. Skipped for PUBLIC PARKS & RECREATION since those are
    # outdoor courts. Does not affect any other sport.
    if (sport_choice == "Basketball"
            and category_name != "PUBLIC PARKS & RECREATION"):
        seen = set()
        for entry in entries:
            key = entry.get("name", "")
            if not key or key in seen:
                continue
            seen.add(key)
            gym_row = _row("Gymnasium Basketball Court",
                            entry["lat"], entry["lon"])
            gym_row["secondary_sport"] = "Volleyball 18U"
            rows.append(gym_row)
    return rows

def build_json(sport_results, city, state):
    # I need to recover
    # FacilityName FacilityCity FacilityCounty FieldName Length Width BaseballLength BaseballWidth FieldType Latitude Longitude ImageLink
    # the other feilds in the JSON are either going to be blank or are unknowable without data we do not have access to here
    # expand_to_rows # use this to extract the data I want

    #sample = "{'FacilityId':'NA','FacilityName':'NA','FacilityCity':'NA','FacilityCounty':'NA','CreatedUser':'NA','CreateDate':'NA'," \
    #"'ReviewedByUser','ReviewedByDate','ReviewCompleted'," \
    #"'FieldId':'NA','FieldName':'NA','Length':'NA','Width':'NA','BaseballLength':'NA','BaseballWidth':'NA','FieldType':'NA','Latitude':'NA'," \
    #"'Longitude':'NA',ImageLink':'NA'}"
    # in the above CreatedUser, ReviewedByUser, ReviewedByDate, ReviewCompleted, FieldId, FacilityId
    
    output_frame = pd.DataFrame()
    section_order = [
            "PUBLIC PARKS & RECREATION",
            "GYMNASIUM / INDOOR FACILITIES",
            "HIGH SCHOOLS",
            "MIDDLE SCHOOLS",
            "ELEMENTARY SCHOOLS",
            "COLLEGE",
            "OTHER FACILITIES",
        ]
    for sport_choice, (categories, total) in sport_results.items():
        if not categories:
            continue
        sport_config = SPORTS_CONFIG[sport_choice]
        for section in section_order:
            entries = categories.get(section, [])
            data_rows = expand_to_rows(entries, sport_config, section, sport_choice)
            df = pd.DataFrame(data_rows)
            df = df.rename(columns=
                           {"lat": "Latitude", "lon": "Longitude",
                            "length_1_ft": "Length", "width_1_ft": "Width",
                            "length_2_ft": "BaseballLength", "width_2_ft": "BaseballWidth",
                            "google_earth_url": "ImageLink", "name": "FacilityName"})
            
            try:
                output_frame = pd.concat([output_frame, df], ignore_index=True)
                print(output_frame)
            except:
                output_frame = df
            #output_json.update(json.load(data_rows))
    output_frame['city'] = city
    output_frame['county'] = ""
    output_frame['state'] = state
    #print(output_json)
    #output_json = json.loads(api_response_string)
    output_frame.to_json('data.json', orient='records', indent=4)
    return output_frame


def build_excel(categories, sport_config, sport_choice, city):
    wb = Workbook()
    ws = wb.active
    title = f"{sport_config['facility_label']}s - {city}"[:31]
    ws.title = title

    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    section_font = Font(name="Arial", size=10, bold=True, color="1F3864")
    section_fill = PatternFill("solid", fgColor="B4C6E7")
    data_font = Font(name="Arial", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    widths = {
        "A": 26, "B": 22, "C": 32, "D": 10, "E": 10,
        "F": 11, "G": 11, "H": 11, "I": 11,
        "J": 14, "K": 10, "L": 26, "M": 38, "N": 9, "O": 14,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    headers = [
        "Name of Facility", "Description", "Address",
        "Lat", "Lon",
        "Length 1 (ft)", "Width 1 (ft)",
        "Length 2 (ft)", "Width 2 (ft)",
        "Primary Sport", "Age Group",
        "Other Primary Sports", "Secondary Sports",
        "ZIP", "Google Earth",
    ]
    NUM_COLS = len(headers)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    current_row = 2
    section_order = [
        "PUBLIC PARKS & RECREATION",
        "GYMNASIUM / INDOOR FACILITIES",
        "HIGH SCHOOLS",
        "MIDDLE SCHOOLS",
        "ELEMENTARY SCHOOLS",
        "COLLEGE",
        "OTHER FACILITIES",
    ]
    total = 0
    link_font = Font(name="Arial", size=10, color="1155CC", underline="single")
    alt_fill = PatternFill("solid", fgColor="F2F2F2")

    for section in section_order:
        entries = categories.get(section, [])
        if not entries:
            continue
        entries.sort(key=lambda x: x["name"].lower())
        data_rows = expand_to_rows(entries, sport_config, section, sport_choice)

        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=NUM_COLS)
        cell = ws.cell(row=current_row, column=1, value=section)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        for c in range(2, NUM_COLS + 1):
            ws.cell(row=current_row, column=c).border = thin_border
            ws.cell(row=current_row, column=c).fill = section_fill
        current_row += 1

        for idx, row in enumerate(data_rows):
            apply_alt = (idx % 2 == 1)
            l1 = row.get("length_1_ft")
            w1 = row.get("width_1_ft")
            l2 = row.get("length_2_ft")
            w2 = row.get("width_2_ft")

            values = [
                row["name"],
                row["description"],
                row["address"],
                round(row["lat"], 4),
                round(row["lon"], 4),
                l1 if l1 else "N/A",
                w1 if w1 else "N/A",
                l2 if l2 else "N/A",
                w2 if w2 else "N/A",
                row.get("primary_sport", ""),
                row.get("age_group", ""),
                row.get("other_primary", ""),
                row.get("secondary_sport", ""),
                row.get("zipcode", ""),
                None,
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col, value=v)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
                if apply_alt:
                    cell.fill = alt_fill

            ge_url = row.get("google_earth_url", "")
            ge_cell = ws.cell(row=current_row, column=NUM_COLS, value="View on Earth")
            ge_cell.hyperlink = ge_url
            ge_cell.font = link_font
            ge_cell.alignment = Alignment(vertical="center", wrap_text=True)
            ge_cell.border = thin_border
            if apply_alt:
                ge_cell.fill = alt_fill

            current_row += 1
            total += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, total

def _read_input_workbook(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)

    cols = {c.lower().strip(): c for c in df.columns}

    def _pick(*candidates):
        for c in candidates:
            if c in cols:
                return cols[c]
        return None

    c_city = _pick("city", "town")
    c_county = _pick("county")
    c_state = _pick("state", "st")
    if not c_city:
        raise ValueError(f"Missing 'City' column. Found: {list(df.columns)}")
    if not c_state:
        raise ValueError(f"Missing 'State' column. Found: {list(df.columns)}")

    rows = []
    for _, r in df.iterrows():
        city = str(r[c_city]).strip()
        county = str(r[c_county]).strip() if c_county and pd.notna(r[c_county]) else ""
        state = str(r[c_state]).strip()
        if not city or city.lower() == "nan":
            continue
        rows.append({"city": city, "county": county, "state": state})
    return rows

def _run_single_job(city, county, state, sport_choice, overpass_url, use_cache,
                     log_lock, status_callback):
    sport_config = SPORTS_CONFIG[sport_choice]
    job_tag = f"[{city}/{sport_choice}]"
    job_logs = []

    def _log(msg):
        job_logs.append(msg)

    try:

        bbox = lookup_city_bbox(city, county, state, "United States",
                                 use_cache=use_cache)
        if not bbox:
            return sport_choice, None, 0, f"{job_tag} bbox lookup failed"

        is_local = ("localhost" in overpass_url or "127.0.0.1" in overpass_url)
        op_results = fetch_overpass(bbox, sport_config, overpass_url, _log,
                                     is_local=is_local, use_cache=use_cache)
        nm_results = fetch_nominatim(city, state, bbox, sport_config, _log,
                                      use_cache=use_cache)
        nces_results = fetch_nces_schools(city, state, bbox, _log,
                                            use_cache=use_cache)

        merged = merge_and_deduplicate(
            op_results + nm_results + nces_results, sport_config, _log)
        if not merged:
            with log_lock:
                status_callback(f"{job_tag} 0 facilities")
            return sport_choice, {}, 0, None

        reverse_geocode_all(merged, city, _log, use_local_nominatim=False,
                             use_cache=use_cache, target_state=state)
        filtered, _ = filter_wrong_city(merged, city, state, bbox, _log,
                                          use_cache=use_cache)
        if not filtered:
            with log_lock:
                status_callback(f"{job_tag} 0 after city filter")
            return sport_choice, {}, 0, None

        categories = categorize(filtered, sport_config, _log)
        total = sum(len(v) for v in categories.values())
        with log_lock:
            status_callback(f"{job_tag} ✅ {total} facilities")
        return sport_choice, categories, total, None
    except Exception as e:
        tb = traceback.format_exc(limit=2)
        with log_lock:
            status_callback(f"{job_tag} ❌ {type(e).__name__}: {e}")
        return sport_choice, None, 0, f"{job_tag} {e}\n{tb}"

def _build_city_workbook(city, sport_results):
    master = Workbook()
    master.remove(master.active)

    grand_total = 0
    for sport_choice, (categories, total) in sport_results.items():
        if not categories:
            continue
        sport_config = SPORTS_CONFIG[sport_choice]
        buf, n = build_excel(categories, sport_config, sport_choice, city)
        grand_total += n

        src_wb = load_workbook(buf)
        src_ws = src_wb.active
        sheet_name = sport_choice.replace(" / ", "_")[:31]
        dst_ws = master.create_sheet(title=sheet_name)
        for row in src_ws.iter_rows(values_only=False):
            for cell in row:
                dst_cell = dst_ws.cell(row=cell.row, column=cell.column,
                                        value=cell.value)
                if cell.has_style:
                    dst_cell.font = cell.font.copy()
                    dst_cell.fill = cell.fill.copy()
                    dst_cell.alignment = cell.alignment.copy()
                    dst_cell.border = cell.border.copy()
                if cell.hyperlink:
                    dst_cell.hyperlink = cell.hyperlink
        for col_letter, dim in src_ws.column_dimensions.items():
            dst_ws.column_dimensions[col_letter].width = dim.width
        dst_ws.freeze_panes = src_ws.freeze_panes
        for merge_range in src_ws.merged_cells.ranges:
            dst_ws.merge_cells(str(merge_range))

    if not master.sheetnames:
        return None, 0

    out = io.BytesIO()
    master.save(out)
    out.seek(0)
    return out, grand_total

def _preresolve_bboxes(rows, use_cache, status_callback):
    status_callback(f"🌐 Pre-resolving bboxes for {len(rows)} cities (serial)...")
    resolved = {}
    failed = []
    for r in rows:
        city = r["city"]
        if city in resolved:
            continue
        bbox = lookup_city_bbox(city, r["county"], r["state"], "United States",
                                 use_cache=use_cache)
        if bbox is None:

            time.sleep(3.0)
            bbox = lookup_city_bbox(city, r["county"], r["state"], "United States",
                                     use_cache=use_cache)
        if bbox is None:
            failed.append(city)
            status_callback(f"  ❌ {city}: bbox lookup failed (after retry)")
        else:
            resolved[city] = bbox
            status_callback(f"  ✅ {city}: bbox resolved")

        time.sleep(1.2)
    return resolved, failed

def run_batch(rows, sports_selected, overpass_url, use_cache, max_workers,
               status_callback):
    log_lock = Lock()

    resolved, failed_cities = _preresolve_bboxes(rows, use_cache, status_callback)

    jobs = []
    for r in rows:
        if r["city"] in failed_cities:
            continue
        for sport in sports_selected:
            jobs.append((r["city"], r["county"], r["state"], sport))

    status_callback(f"🚀 Submitting {len(jobs)} jobs to {max_workers} workers "
                    f"({len(resolved)} cities × {len(sports_selected)} sports)")

    results = {r["city"]: {} for r in rows}
    errors = []
    done = 0
    t0 = time.time()

    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {
            ex.submit(_run_single_job, c, co, s, sp, overpass_url, use_cache,
                       log_lock, status_callback): (c, sp)
            for c, co, s, sp in jobs
        }
        for fut in as_completed(futures):
            city, sport = futures[fut]
            try:
                sp, cats, total, err = fut.result()
                if err:
                    errors.append(err)
                if cats is not None:
                    results[city][sp] = (cats, total)
            except Exception as e:
                errors.append(f"[{city}/{sport}] worker crashed: {e}")
            done += 1
            with log_lock:
                status_callback(f"  Progress: {done}/{len(jobs)}  "
                                f"({(time.time() - t0):.1f}s elapsed)")

    for fc in failed_cities:
        errors.append(f"[{fc}] bbox lookup failed (skipped all sports)")

    return results, errors

def _build_zip(per_city_results, status_callback):
    #print(per_city_results)
    buf = io.BytesIO()
    grand_total = 0
    cities_with_data = 0
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for city, sport_results in per_city_results.items():
            if not sport_results:
                continue
            wb_buf, n = _build_city_workbook(city, sport_results)
            if wb_buf is None:
                continue
            safe_name = city.replace(" ", "_").replace("/", "_")
            zf.writestr(f"{safe_name}.xlsx", wb_buf.getvalue())
            cities_with_data += 1
            grand_total += n
            status_callback(f"  📦 Packed {safe_name}.xlsx ({n} rows)")
    buf.seek(0)
    return buf, cities_with_data, grand_total

def main():
    st.set_page_config(
        page_title="Batch Sports Facility Finder",
        page_icon="🏟️",
        layout="wide",
    )
    st.title("🏟️ Batch Sports Facility Finder")
    st.markdown("""
    Upload a spreadsheet of cities and get one Excel workbook per city —
    each containing a sheet per sport (Soccer, Baseball, Basketball, Tennis,
    Volleyball). Runs all jobs in parallel.
    """)

    with st.sidebar:
        st.header("⚙️ Settings")
        sports_selected = st.multiselect(
            "Sports to fetch",
            options=list(SPORTS_CONFIG.keys()),
            default=list(SPORTS_CONFIG.keys()),
        )

        mirror_options = ["Auto (try all public mirrors)"] + OVERPASS_MIRRORS + ["Custom URL..."]
        mirror_choice = st.selectbox("Overpass endpoint", mirror_options, index=0)
        if mirror_choice == "Custom URL...":
            overpass_url = st.text_input(
                "Custom Overpass URL",
                value=os.environ.get("OVERPASS_URL", "http://localhost:8080/api/interpreter"),
            )
        elif mirror_choice == "Auto (try all public mirrors)":
            overpass_url = OVERPASS_MIRRORS[0]
        else:
            overpass_url = mirror_choice

        is_local = ("localhost" in overpass_url or "127.0.0.1" in overpass_url)
        default_workers = 8 if is_local else 3
        max_workers = st.slider(
            "Parallel workers",
            min_value=1, max_value=16, value=default_workers,
            help=("Threads running (city × sport) jobs concurrently. Public "
                  "Overpass + Nominatim throttle aggressively — keep ≤3 for "
                  "public mirrors. Use 8+ with a local Overpass."),
        )
        use_cache = st.checkbox("Use response cache", value=True)

        count, size = cache_stats()
        if count > 0:
            st.caption(f"💾 Cache: {count} entries, {size/1024/1024:.1f} MB")
            if st.button("🗑️ Clear cache", use_container_width=True):
                cache_clear()
                st.success("Cleared.")
        else:
            st.caption("💾 Cache empty")

        st.divider()
        st.subheader("🔍 Single city search")
        single_city = st.text_input("City", key="single_city",
                                     placeholder="Daly City")
        single_county = st.text_input("County (optional)", key="single_county",
                                       placeholder="San Mateo County")
        single_state = st.text_input("State", key="single_state",
                                      placeholder="California")
        run_single = st.button("🔎 Search this city", use_container_width=True,
                                type="primary")

    if run_single:
        if not single_city.strip() or not single_state.strip():
            st.error("City and State are required for single-city search.")
            return
        if not sports_selected:
            st.error("Pick at least one sport in the sidebar.")
            return
        row = {
            "city": single_city.strip(),
            "county": (single_county or "").strip(),
            "state": single_state.strip(),
        }
        st.subheader(f"🏟️ {row['city']}, {row['state']}")
        st.info(f"Running {len(sports_selected)} sport(s) on "
                f"**{max_workers}** workers.")
        log_messages = []
        def _log(msg):
            log_messages.append(msg)
        status = st.status(f"Fetching {row['city']}...", expanded=True)
        with status:
            t0 = time.time()
            per_city_results, errors = run_batch(
                [row], sports_selected, overpass_url, use_cache,
                max_workers, _log,
            )
            elapsed = time.time() - t0
            sport_results = per_city_results.get(row["city"], {})
            build_json(sport_results, row['city'], row['state'])
            wb_buf, n = _build_city_workbook(row["city"], sport_results) \
                if sport_results else (None, 0)
            status.update(
                label=f"✅ Done in {elapsed:.1f}s — {n} facility rows",
                state="complete",
            )
        if wb_buf is None or n == 0:
            st.warning("No facilities found for this city.")
        else:
            st.success(f"✅ {n} facility rows across "
                       f"{len(sport_results)} sport(s)")
            for sp, (_cats, tot) in sport_results.items():
                st.caption(f"• {sp}: {tot}")
            safe_name = row["city"].replace(" ", "_").replace("/", "_")
            st.download_button(
                label=f"📥 Download {safe_name}.xlsx",
                data=wb_buf,
                file_name=f"{safe_name}.xlsx",
                mime=("application/vnd.openxmlformats-officedocument."
                      "spreadsheetml.sheet"),
                type="primary",
                use_container_width=True,
            )
        if errors:
            with st.expander(f"⚠️ {len(errors)} error(s)"):
                for e in errors:
                    st.text(e)
        with st.expander(f"📜 Log ({len(log_messages)} lines)"):
            st.code("\n".join(log_messages), language="text")
        return

    st.subheader("1. Upload city list")
    st.caption("Required columns: **City**, **State**. Optional: **County**. "
                "Accepts .xlsx or .csv.")

    sample = pd.DataFrame([
        {"City": "Daly City", "County": "San Mateo County", "State": "California"},
        {"City": "Berkeley", "County": "Alameda County", "State": "California"},
    ])
    with st.expander("📋 Sample input"):
        st.dataframe(sample, use_container_width=True, hide_index=True)

    uploaded = st.file_uploader("Upload spreadsheet",
                                  type=["xlsx", "xls", "csv"])
    if not uploaded:
        st.info("👆 Upload a city list to begin.")
        return

    try:
        rows = _read_input_workbook(uploaded)
    except Exception as e:
        st.error(f"Failed to parse upload: {e}")
        return

    if not rows:
        st.warning("No valid rows found in upload.")
        return

    st.success(f"Parsed {len(rows)} cities")
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    if not sports_selected:
        st.warning("Pick at least one sport in the sidebar.")
        return

    total_jobs = len(rows) * len(sports_selected)
    st.info(f"Will run **{total_jobs}** jobs ({len(rows)} cities × "
             f"{len(sports_selected)} sports) on **{max_workers}** workers.")

    if not st.button("🚀 Run Batch", type="primary", use_container_width=True):
        return

    log_messages = []

    def log(msg):
        log_messages.append(msg)

    status = st.status(f"Running {total_jobs} jobs in parallel...",
                        expanded=True)
    with status:
        t0 = time.time()
        per_city_results, errors = run_batch(rows, sports_selected,
                                               overpass_url, use_cache,
                                               max_workers, log)
        st.write(f"All jobs done in {time.time()-t0:.1f}s. Building workbooks...")
        zip_buf, cities_done, grand_total = _build_zip(per_city_results, log)
        status.update(label=f"✅ Done — {cities_done} cities, "
                              f"{grand_total} facility rows",
                        state="complete")

    st.success(f"✅ Built {cities_done} workbooks with {grand_total} total rows")

    if errors:
        with st.expander(f"⚠️ {len(errors)} job error(s)"):
            for e in errors:
                st.text(e)

    with st.expander(f"📜 Full log ({len(log_messages)} lines)"):
        st.code("\n".join(log_messages), language="text")

    st.download_button(
        label="📥 Download results.zip",
        data=zip_buf,
        file_name=f"facilities_batch_{int(time.time())}.zip",
        mime="application/zip",
        type="primary",
        use_container_width=True,
    )

if __name__ == "__main__":
    main()
